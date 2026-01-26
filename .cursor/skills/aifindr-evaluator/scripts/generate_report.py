#!/usr/bin/env python3
"""
Generate evaluation reports (Markdown + XLSX) from evaluation results.

Usage:
    python generate_report.py --output-dir evals/project/runs/2026-01-21_14-30/ --results results.json

Or use programmatically:
    from generate_report import generate_reports
    generate_reports(results, output_dir)
"""

import argparse
import json
import os
from datetime import datetime
from pathlib import Path
from typing import Any

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# =============================================================================
# XLSX Styling Constants (deterministic look and feel)
# =============================================================================

COLORS = {
    "header_bg": "2F5496",
    "header_font": "FFFFFF",
    "pass_bg": "C6EFCE",
    "pass_font": "006100",
    "fail_bg": "FFC7CE",
    "fail_font": "9C0006",
    "partial_bg": "FFEB9C",
    "partial_font": "9C5700",
    "no_retrieval_bg": "D9D9D9",
    "no_retrieval_font": "595959",
    "alt_row_bg": "F2F2F2",
    "border": "B4B4B4",
}

COLUMN_WIDTHS = {
    "id": 6,
    "query": 40,
    "expected": 50,
    "response": 60,
    "verdict": 12,
    "num_sources": 12,
    "latency_s": 10,
    "notes": 45,
}

HEADERS = ["id", "query", "expected", "response", "verdict", "num_sources", "latency_s", "notes"]


def get_verdict_style(verdict: str) -> tuple:
    """Return (fill, font) for a verdict."""
    verdict_upper = verdict.upper() if verdict else ""
    if verdict_upper == "PASS":
        return (
            PatternFill(start_color=COLORS["pass_bg"], end_color=COLORS["pass_bg"], fill_type="solid"),
            Font(color=COLORS["pass_font"], bold=True)
        )
    elif verdict_upper == "FAIL":
        return (
            PatternFill(start_color=COLORS["fail_bg"], end_color=COLORS["fail_bg"], fill_type="solid"),
            Font(color=COLORS["fail_font"], bold=True)
        )
    elif verdict_upper == "PARTIAL":
        return (
            PatternFill(start_color=COLORS["partial_bg"], end_color=COLORS["partial_bg"], fill_type="solid"),
            Font(color=COLORS["partial_font"], bold=True)
        )
    else:  # NO_RETRIEVAL or unknown
        return (
            PatternFill(start_color=COLORS["no_retrieval_bg"], end_color=COLORS["no_retrieval_bg"], fill_type="solid"),
            Font(color=COLORS["no_retrieval_font"], bold=True)
        )


def create_xlsx_report(results: list[dict], output_path: str, metadata: dict = None) -> str:
    """
    Create an XLSX report with consistent styling.

    Args:
        results: List of evaluation result dicts with keys matching HEADERS
        output_path: Path to save the XLSX file
        metadata: Optional metadata dict with project, date, etc.

    Returns:
        Path to the created file
    """
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl is required. Install with: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Evaluation Results"

    # Define styles
    header_fill = PatternFill(start_color=COLORS["header_bg"], end_color=COLORS["header_bg"], fill_type="solid")
    header_font = Font(bold=True, color=COLORS["header_font"], size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    cell_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    center_alignment = Alignment(horizontal="center", vertical="top")

    thin_border = Border(
        left=Side(style='thin', color=COLORS["border"]),
        right=Side(style='thin', color=COLORS["border"]),
        top=Side(style='thin', color=COLORS["border"]),
        bottom=Side(style='thin', color=COLORS["border"])
    )

    alt_row_fill = PatternFill(start_color=COLORS["alt_row_bg"], end_color=COLORS["alt_row_bg"], fill_type="solid")

    # Write headers
    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header.upper())
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # Freeze header row
    ws.freeze_panes = "A2"

    # Write data rows
    for row_idx, result in enumerate(results, 2):
        is_alt_row = row_idx % 2 == 0

        for col_idx, header in enumerate(HEADERS, 1):
            value = result.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

            # Apply alignment
            if header in ["id", "verdict", "num_sources", "latency_s"]:
                cell.alignment = center_alignment
            else:
                cell.alignment = cell_alignment

            # Apply verdict-specific styling
            if header == "verdict":
                fill, font = get_verdict_style(str(value))
                cell.fill = fill
                cell.font = font
            elif is_alt_row:
                cell.fill = alt_row_fill

    # Set column widths
    for col_idx, header in enumerate(HEADERS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COLUMN_WIDTHS.get(header, 15)

    # Set row heights (estimate based on content)
    ws.row_dimensions[1].height = 25  # Header row
    for row_idx in range(2, len(results) + 2):
        # Calculate height based on longest cell content
        max_lines = 1
        for col_idx, header in enumerate(HEADERS, 1):
            value = str(results[row_idx - 2].get(header, ""))
            width = COLUMN_WIDTHS.get(header, 15)
            estimated_lines = max(1, len(value) // (width * 1.5) + value.count('\n') + 1)
            max_lines = max(max_lines, estimated_lines)
        ws.row_dimensions[row_idx].height = min(400, max(30, max_lines * 15))

    # Add autofilter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(results) + 1}"

    # Add Summary sheet
    ws_summary = wb.create_sheet("Summary", 0)
    ws_summary.sheet_properties.tabColor = "2F5496"

    # Summary content
    summary_data = [
        ("EVALUATION SUMMARY", ""),
        ("", ""),
        ("Project", metadata.get("project", "N/A") if metadata else "N/A"),
        ("Date", metadata.get("date", datetime.now().strftime("%Y-%m-%d %H:%M")) if metadata else datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Total Queries", len(results)),
        ("", ""),
        ("VERDICTS", "COUNT"),
    ]

    # Count verdicts
    verdict_counts = {"PASS": 0, "PARTIAL": 0, "FAIL": 0, "NO_RETRIEVAL": 0}
    for r in results:
        v = str(r.get("verdict", "")).upper()
        if v in verdict_counts:
            verdict_counts[v] += 1

    for verdict, count in verdict_counts.items():
        pct = f"{count / len(results) * 100:.1f}%" if results else "0%"
        summary_data.append((verdict, f"{count} ({pct})"))

    # Write summary
    for row_idx, (label, value) in enumerate(summary_data, 1):
        cell_label = ws_summary.cell(row=row_idx, column=1, value=label)
        cell_value = ws_summary.cell(row=row_idx, column=2, value=value)

        if row_idx == 1:
            cell_label.font = Font(bold=True, size=14, color=COLORS["header_bg"])
        elif label in ["VERDICTS", "Project", "Date", "Total Queries"]:
            cell_label.font = Font(bold=True)
        elif label in verdict_counts:
            fill, font = get_verdict_style(label)
            cell_label.fill = fill
            cell_label.font = font

    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 25

    # Save
    wb.save(output_path)
    return output_path


def create_markdown_report(results: list[dict], output_path: str, metadata: dict = None) -> str:
    """
    Create a Markdown report.

    Args:
        results: List of evaluation result dicts
        output_path: Path to save the markdown file
        metadata: Optional metadata dict

    Returns:
        Path to the created file
    """
    meta = metadata or {}
    project = meta.get("project", "N/A")
    date = meta.get("date", datetime.now().strftime("%Y-%m-%d %H:%M"))
    kb_path = meta.get("knowledge_base", "N/A")

    # Count verdicts
    verdict_counts = {"PASS": 0, "PARTIAL": 0, "FAIL": 0, "NO_RETRIEVAL": 0}
    for r in results:
        v = str(r.get("verdict", "")).upper()
        if v in verdict_counts:
            verdict_counts[v] += 1

    total = len(results)

    lines = [
        "# Evaluation Report",
        "",
        f"**Project:** {project}",
        f"**Date:** {date}",
        f"**Queries evaluated:** {total}",
        f"**Knowledge base:** {kb_path}",
        "",
        "## Summary",
        "",
        "| Verdict | Count | Percentage |",
        "|---------|-------|------------|",
    ]

    for verdict, count in verdict_counts.items():
        pct = f"{count / total * 100:.0f}%" if total else "0%"
        lines.append(f"| {verdict} | {count} | {pct} |")

    lines.extend([
        "",
        "## Results",
        "",
        "| # | Query | Verdict | Latency | Notes |",
        "|---|-------|---------|---------|-------|",
    ])

    for r in results:
        rid = r.get("id", "")
        query = r.get("query", "")[:50] + ("..." if len(r.get("query", "")) > 50 else "")
        verdict = r.get("verdict", "")
        latency = f"{r.get('latency_s', 0):.1f}s"
        notes = r.get("notes", "")[:60] + ("..." if len(r.get("notes", "")) > 60 else "")
        lines.append(f"| {rid} | {query} | {verdict} | {latency} | {notes} |")

    # Detailed results
    lines.extend([
        "",
        "## Detailed Results",
        "",
    ])

    for r in results:
        rid = r.get("id", "")
        query = r.get("query", "")
        verdict = r.get("verdict", "")
        response = r.get("response", "")
        expected = r.get("expected", "")
        notes = r.get("notes", "")

        verdict_emoji = {"PASS": "✅", "FAIL": "❌", "PARTIAL": "⚠️", "NO_RETRIEVAL": "🔍"}.get(verdict.upper(), "❓")

        lines.extend([
            f"### Query {rid}: {query}",
            "",
            f"**Verdict:** {verdict} {verdict_emoji}",
            "",
            "**Agent Response:**",
            f"> {response.replace(chr(10), chr(10) + '> ')}",
            "",
        ])

        if expected:
            lines.extend([
                "**Expected:**",
                f"> {expected.replace(chr(10), chr(10) + '> ')}",
                "",
            ])

        if notes:
            lines.extend([
                f"**Notes:** {notes}",
                "",
            ])

        lines.append("---")
        lines.append("")

    # Issues and recommendations
    fails = [r for r in results if r.get("verdict", "").upper() in ["FAIL", "NO_RETRIEVAL"]]

    lines.extend([
        "## Issues Found",
        "",
    ])

    if fails:
        for r in fails:
            lines.append(f"- **Query {r.get('id')}** ({r.get('verdict')}): {r.get('notes', 'No notes')}")
    else:
        lines.append("*No issues found in this evaluation.*")

    lines.extend([
        "",
        "## Recommendations",
        "",
    ])

    if fails:
        lines.append("Review the failed queries above and consider:")
        lines.append("- Adjusting chunking strategy if retrieval is failing")
        lines.append("- Adding Q&A entries for terminology gaps")
        lines.append("- Reviewing prompt instructions if responses are incorrect")
    else:
        lines.append("*No recommendations - all queries passed.*")

    content = "\n".join(lines)

    Path(output_path).write_text(content, encoding="utf-8")
    return output_path


def generate_reports(
    results: list[dict],
    output_dir: str,
    metadata: dict = None
) -> tuple[str, str]:
    """
    Generate both Markdown and XLSX reports.

    Args:
        results: List of evaluation result dicts
        output_dir: Directory to save reports
        metadata: Optional metadata dict

    Returns:
        Tuple of (markdown_path, xlsx_path)
    """
    Path(output_dir).mkdir(parents=True, exist_ok=True)

    md_path = os.path.join(output_dir, "report.md")
    xlsx_path = os.path.join(output_dir, "results.xlsx")

    create_markdown_report(results, md_path, metadata)
    create_xlsx_report(results, xlsx_path, metadata)

    return md_path, xlsx_path


def get_run_folder_name() -> str:
    """Get a run folder name with date and time."""
    return datetime.now().strftime("%Y-%m-%d_%H-%M-%S")


def main():
    parser = argparse.ArgumentParser(description="Generate evaluation reports")
    parser.add_argument("--results", "-r", required=True, help="Path to results JSON file")
    parser.add_argument("--output-dir", "-o", required=True, help="Output directory for reports")
    parser.add_argument("--project", "-p", default="", help="Project ID for metadata")
    parser.add_argument("--knowledge-base", "-k", default="", help="Knowledge base path for metadata")
    args = parser.parse_args()

    # Load results
    with open(args.results, "r", encoding="utf-8") as f:
        results = json.load(f)

    if not isinstance(results, list):
        results = [results]

    metadata = {
        "project": args.project,
        "knowledge_base": args.knowledge_base,
        "date": datetime.now().strftime("%Y-%m-%d %H:%M"),
    }

    md_path, xlsx_path = generate_reports(results, args.output_dir, metadata)

    print(f"Reports generated:")
    print(f"  Markdown: {md_path}")
    print(f"  XLSX: {xlsx_path}")


if __name__ == "__main__":
    main()
